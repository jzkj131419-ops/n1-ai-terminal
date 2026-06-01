#!/usr/bin/env python3
"""
广发信用卡差异错因分析
目标：把577缺失行+398多出行 按假说归因，找出错因模型
输出：TASK_广发差异错因分析_核查清单.xlsx
"""
import openpyxl
from openpyxl.styles import PatternFill, Font
from collections import Counter, defaultdict
from datetime import datetime, date, timedelta
from pathlib import Path

SRC_XLSX  = Path("/Volumes/AI审计/周贤查账/广发/55776信用卡流水.xlsx")
ATT_XLSX  = Path("/Volumes/AI审计/AI案件汇总_周贤_20260531_民生修正版/01_广发/周贤_审计附表.xlsx")
OUT_XLSX  = Path("/Volumes/AI审计/AI案件汇总_周贤_20260531_民生修正版/01_广发/广发_差异错因核查清单.xlsx")

RED    = PatternFill("solid", fgColor="FF9999")  # 真实漏录（无法解释）
AMBER  = PatternFill("solid", fgColor="FFD580")  # 可解释差异
GREEN  = PatternFill("solid", fgColor="C6EFCE")  # 多出行
BLUE   = PatternFill("solid", fgColor="BDD7EE")  # 还款/个人/周期（重点关注）

def to_ymd(v):
    if isinstance(v, (datetime, date)):
        return v.date() if isinstance(v, datetime) else v
    s = str(v).replace('-','')
    return date(int(s[:4]), int(s[4:6]), int(s[6:8]))

# ── 读源 Excel ────────────────────────────────────────────────
print("读取源 Excel...")
wb_s = openpyxl.load_workbook(SRC_XLSX)
ws_s = wb_s['信用卡交易流水(1)']
src_rows = []
for r in ws_s.iter_rows(min_row=2, values_only=True):
    if r[1] != '5203821372910250': continue
    if not r[10] or float(r[10]) == 0: continue
    src_rows.append({
        'date': to_ymd(r[3]),
        'amt':  round(float(r[10]), 2),
        'code': str(r[4] or ''),
        'code_cn': str(r[5] or ''),
        'desc': str(r[12] or '').strip(),
        'raw': r,
    })

# ── 读附表 ────────────────────────────────────────────────────
print("读取附表...")
wb_a = openpyxl.load_workbook(ATT_XLSX)
ws_a = wb_a['1交易记录汇总']
att_rows = []
for r in ws_a.iter_rows(min_row=2, values_only=True):
    if not r[0]: continue
    att_rows.append({
        'date': to_ymd(r[0]),
        'amt':  round(float(r[6]) if r[6] else 0, 2),
        'desc': str(r[13] or '').strip(),
        'cp':   str(r[11] or '').strip(),
        'raw': r,
    })

# ── 计算双向差异 ──────────────────────────────────────────────
src_counter = Counter((r['date'], r['amt']) for r in src_rows)
att_counter = Counter((r['date'], r['amt']) for r in att_rows)
missing_c = src_counter - att_counter   # 源有附表无
extra_c   = att_counter - src_counter   # 附表有源无

print(f"缺失: {sum(missing_c.values())} 行  多出: {sum(extra_c.values())} 行  净差: {sum(missing_c.values())-sum(extra_c.values())} 行")

# ── 构建附表按日期的快速查找 ─────────────────────────────────
att_by_date = defaultdict(list)
for r in att_rows:
    att_by_date[r['date']].append(r)

# ── 归因函数 ─────────────────────────────────────────────────
REPAYMENT_KEYWORDS = ['还款', '跨行清算', '人行', '还信用卡']
PERSONAL_KEYWORDS  = ['孙志辉', '张改', '赵常清', '邵芳婷', '张苗苗']
PERIODIC_KEYWORDS  = ['首开千方', '停车', '物业', '公积金', '社保']
REFUND_CODES       = {'1006'}  # 退货

def classify(row):
    desc = row['desc']
    code = row['code']
    if any(k in desc for k in REPAYMENT_KEYWORDS):
        return '还款类', '重点：资金来源待核查', BLUE
    if any(k in desc for k in PERSONAL_KEYWORDS):
        return '个人转账类', '重点：个人转账红旗', BLUE
    if any(k in desc for k in PERIODIC_KEYWORDS):
        return '周期性支出类', '规律性消费，核实性质', AMBER
    if code in REFUND_CODES:
        return '退货冲正类', '交易码1006退货，可能被解析为正向', AMBER
    return '疑似错位/精度', '金额精度或数量配对问题', AMBER

def find_near_match(row, tolerance_days=15, tolerance_pct=0.05):
    """在附表中找近似匹配（日期偏移或金额精度差异）"""
    target_date = row['date']
    target_amt  = row['amt']
    best = None
    best_score = 999999
    for d in att_by_date:
        if abs((d - target_date).days) > tolerance_days:
            continue
        for a_row in att_by_date[d]:
            amt_diff = abs(a_row['amt'] - target_amt)
            if target_amt > 0 and amt_diff / target_amt > tolerance_pct:
                continue
            score = abs((d - target_date).days) * 10 + amt_diff
            if score < best_score:
                best_score = score
                best = (a_row, abs((d - target_date).days), round(amt_diff, 2))
    return best

# ── 展开缺失行 ────────────────────────────────────────────────
print("归因分析中...")
missing_items = []
for (d, a), cnt in sorted(missing_c.items()):
    # 找源数据里的描述
    descs = [r['desc'] for r in src_rows if r['date']==d and r['amt']==a]
    codes = [r['code'] for r in src_rows if r['date']==d and r['amt']==a]
    for i in range(cnt):
        desc = descs[i] if i < len(descs) else ''
        code = codes[i] if i < len(codes) else ''
        row_dummy = {'date': d, 'amt': a, 'desc': desc, 'code': code}
        cat, reason, fill = classify(row_dummy)
        near = find_near_match(row_dummy)
        missing_items.append({
            'date': d, 'amt': a, 'desc': desc, 'code': code,
            'category': cat, 'reason': reason, 'fill': fill,
            'near_match_date': near[0]['date'] if near else '',
            'near_match_amt':  near[0]['amt']  if near else '',
            'date_diff':       near[1]          if near else '',
            'amt_diff':        near[2]          if near else '',
        })

# ── 展开多出行 ────────────────────────────────────────────────
extra_items = []
for (d, a), cnt in sorted(extra_c.items()):
    descs = [r['desc'] for r in att_rows if r['date']==d and r['amt']==a]
    for i in range(cnt):
        desc = descs[i] if i < len(descs) else ''
        extra_items.append({'date': d, 'amt': a, 'desc': desc})

# ── 写核查清单 ────────────────────────────────────────────────
print("写核查清单...")
out_wb = openpyxl.Workbook()

# Sheet1：缺失行归因
ws1 = out_wb.active
ws1.title = "缺失577行（源有附表无）"
headers1 = ['日期','金额','交易描述','交易码','错因分类','说明',
            '近似匹配日期','近似匹配金额','日期差(天)','金额差','人工结论']
ws1.append(headers1)
for cell in ws1[1]:
    cell.font = Font(bold=True)

for item in missing_items:
    row = [item['date'], item['amt'], item['desc'], item['code'],
           item['category'], item['reason'],
           item['near_match_date'], item['near_match_amt'],
           item['date_diff'], item['amt_diff'], '']
    ws1.append(row)
    ri = ws1.max_row
    for ci in range(1, len(headers1)+1):
        ws1.cell(ri, ci).fill = item['fill']

# Sheet2：多出行
ws2 = out_wb.create_sheet("多出398行（附表有源无）")
headers2 = ['日期','金额','附表描述','疑似原因','人工结论']
ws2.append(headers2)
for cell in ws2[1]:
    cell.font = Font(bold=True)
for item in extra_items:
    # 多出行：找源里有没有近似匹配
    near_src = None
    for r in src_rows:
        if abs((r['date'] - item['date']).days) <= 15 and abs(r['amt'] - item['amt'])/max(item['amt'],0.01) < 0.05:
            near_src = r
            break
    reason = '可能是金额四舍五入错配' if near_src else '无近似源数据，疑似重复录入'
    ws2.append([item['date'], item['amt'], item['desc'], reason, ''])
    ri = ws2.max_row
    fill = AMBER if near_src else RED
    for ci in range(1, len(headers2)+1):
        ws2.cell(ri, ci).fill = fill

# Sheet3：统计汇总
ws3 = out_wb.create_sheet("错因统计")
from collections import Counter as C2
cat_count = C2(item['category'] for item in missing_items)
ws3.append(['错因分类','行数','占比','含义'])
total = sum(cat_count.values())
for cat, n in cat_count.most_common():
    ws3.append([cat, n, f"{n/total*100:.1f}%", ''])

ws3.append([])
ws3.append(['缺失577行金额合计', sum(i['amt'] for i in missing_items)])
ws3.append(['多出398行金额合计', sum(i['amt'] for i in extra_items)])
ws3.append(['净差179行净金额', sum(i['amt'] for i in missing_items) - sum(i['amt'] for i in extra_items)])
ws3.append([])
ws3.append(['蓝色=重点关注（还款/个人转账）', sum(1 for i in missing_items if i['fill']==BLUE)])
ws3.append(['橙色=可解释差异', sum(1 for i in missing_items if i['fill']==AMBER)])

out_wb.save(OUT_XLSX)
print(f"\n核查清单已生成: {OUT_XLSX}")

# ── 打印摘要 ──────────────────────────────────────────────────
print("\n=== 错因分布 ===")
for cat, n in cat_count.most_common():
    print(f"  {cat}: {n} 行")
print(f"\n重点关注（蓝色）: {sum(1 for i in missing_items if i['fill']==BLUE)} 行")
blue_items = [i for i in missing_items if i['fill']==BLUE]
for i in blue_items:
    print(f"  {i['date']} ¥{i['amt']:,.2f} {i['desc'][:30]}")
